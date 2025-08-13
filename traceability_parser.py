# -*- coding: utf-8 -*-
import pandas as pd
from pathlib import Path
from typing import List, Optional
import numpy as np
import sys
from traceability_model import TraceabilityItem, TraceabilityDatabase

# Ensure UTF-8 encoding for Windows compatibility
if sys.platform == 'win32':
    import locale
    try:
        locale.setlocale(locale.LC_ALL, 'sv_SE.UTF-8')
    except locale.Error:
        try:
            locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
        except locale.Error:
            pass  # Use system default


class TraceabilityParser:
    def __init__(self):
        self.database = TraceabilityDatabase()
    
    def parse_file(self, file_path: Path, original_name: str = None) -> List[TraceabilityItem]:
        if file_path.suffix.lower() not in ['.xlsx', '.xls']:
            raise ValueError(f"Unsupported file type: {file_path.suffix}")
        
        # Use original name if provided, otherwise use file path name
        file_name = (original_name or file_path.name).lower()
        
        if 'sök i spårbarhet' in file_name:
            return self._parse_search_file(file_path)
        elif 'lagerlogg' in file_name:
            return self._parse_lagerlogg_file(file_path)
        elif 'nivålista' in file_name:
            return self._parse_nivalista_file(file_path)
        else:
            # Try to detect type by content
            return self._parse_generic_file(file_path)
    
    def _parse_search_file(self, file_path: Path) -> List[TraceabilityItem]:
        items = []
        try:
            df = pd.read_excel(file_path, engine='openpyxl' if file_path.suffix == '.xlsx' else 'xlrd')
            
            # Find relevant columns
            columns = df.columns.tolist()
            artikelnummer_col = self._find_column(columns, ['artikelnummer', 'artikel', 'art.nr'])
            batch_col = self._find_column(columns, ['serienummer/batchnummer', 'batchnummer', 'batch', 'serienummer'])
            charge_col = self._find_column(columns, ['chargenummer', 'charge', 'chargenr'])
            benaming_col = self._find_column(columns, ['artikelbenämning', 'benämning', 'beskrivning'])
            order_col = self._find_column(columns, ['order', 'ordernummer', 'ordernr'])
            
            for _, row in df.iterrows():
                if artikelnummer_col and pd.notna(row[artikelnummer_col]):
                    item = TraceabilityItem(
                        artikelnummer=str(row[artikelnummer_col]),
                        artikelbenaming=str(row[benaming_col]) if benaming_col and pd.notna(row[benaming_col]) else None,
                        batchnummer=str(row[batch_col]) if batch_col and pd.notna(row[batch_col]) else None,
                        chargenummer=str(row[charge_col]) if charge_col and pd.notna(row[charge_col]) else None,
                        serienummer=str(row[batch_col]) if batch_col and pd.notna(row[batch_col]) else None,
                        ordernummer=str(row[order_col]) if order_col and pd.notna(row[order_col]) else None,
                        source_file=str(file_path),
                        source_type='spårbarhet'
                    )
                    items.append(item)
                    self.database.add_item(item)
        except Exception as e:
            print(f"Error parsing {file_path}: {e}")
        
        return items
    
    def _parse_lagerlogg_file(self, file_path: Path) -> List[TraceabilityItem]:
        items = []
        try:
            df = pd.read_excel(file_path, engine='openpyxl' if file_path.suffix == '.xlsx' else 'xlrd')
            
            # Find relevant columns
            columns = df.columns.tolist()
            artikelnummer_col = self._find_column(columns, ['artikelnummer', 'artikel', 'art.nr'])
            batch_col = self._find_column(columns, ['batchnummer', 'batch', 'batchnr'])
            charge_col = self._find_column(columns, ['chargenummer', 'charge', 'chargenr'])
            benaming_col = self._find_column(columns, ['artikelbenämning', 'benämning', 'beskrivning'])
            order_col = self._find_column(columns, ['ordernummer', 'order', 'ordn.nr'])
            
            for _, row in df.iterrows():
                if artikelnummer_col and pd.notna(row[artikelnummer_col]):
                    item = TraceabilityItem(
                        artikelnummer=str(row[artikelnummer_col]),
                        artikelbenaming=str(row[benaming_col]) if benaming_col and pd.notna(row[benaming_col]) else None,
                        batchnummer=str(row[batch_col]) if batch_col and pd.notna(row[batch_col]) else None,
                        chargenummer=str(row[charge_col]) if charge_col and pd.notna(row[charge_col]) else None,
                        ordernummer=str(row[order_col]) if order_col and pd.notna(row[order_col]) else None,
                        source_file=str(file_path),
                        source_type='lagerlogg'
                    )
                    items.append(item)
                    self.database.add_item(item)
        except Exception as e:
            print(f"Error parsing {file_path}: {e}")
        
        return items
    
    def _parse_nivalista_file(self, file_path: Path) -> List[TraceabilityItem]:
        items = []
        try:
            # Use openpyxl to read both data and outline structure
            from openpyxl import load_workbook
            
            # Load workbook to get outline information
            wb = load_workbook(file_path, read_only=False)
            ws = wb.active
            
            # Check if there are any grouped/hidden rows
            print(f"Checking for grouped rows...")
            has_groups = False
            for row_num in range(2, ws.max_row + 1):
                row_dim = ws.row_dimensions.get(row_num)
                if row_dim and (row_dim.hidden or row_dim.outline_level > 0):
                    has_groups = True
                    print(f"Row {row_num}: hidden={row_dim.hidden if row_dim else False}, outline_level={row_dim.outline_level if row_dim else 0}")
            
            if not has_groups:
                print("No grouped rows found - file may not have outline structure")
            
            # Unhide all rows to ensure we get everything
            for row_num in range(1, ws.max_row + 1):
                row_dim = ws.row_dimensions.get(row_num)
                if row_dim:
                    row_dim.hidden = False
            
            # First check the actual Excel structure
            print(f"\nAnalyzing Excel structure:")
            max_row = ws.max_row
            print(f"Excel max_row: {max_row}")
            
            # Count actual data rows and check outline levels
            data_rows = 0
            outline_levels_found = []
            for row in range(2, max_row + 1):  # Skip header
                artikel_cell = ws.cell(row=row, column=3)  # Column C
                if artikel_cell.value is not None:
                    data_rows += 1
                    row_dim = ws.row_dimensions.get(row)
                    outline_level = row_dim.outline_level if row_dim else 0
                    outline_levels_found.append(outline_level)
                    
            print(f"Actual data rows in Excel: {data_rows}")
            print(f"Outline levels found: {set(outline_levels_found)}")
            print(f"Max outline level: {max(outline_levels_found) if outline_levels_found else 0}")
            
            # Also load with pandas for easy data access
            df = pd.read_excel(file_path, engine='openpyxl' if file_path.suffix == '.xlsx' else 'xlrd')
            
            # Find relevant columns
            columns = df.columns.tolist()
            artikelnummer_col = self._find_column(columns, ['artikel/operation', 'artikelnummer', 'artikel'])
            benaming_col = self._find_column(columns, ['benämning', 'beskrivning'])
            artikeltyp_col = self._find_column(columns, ['artikeltyp/operation', 'artikeltyp'])
            kvantitet_col = self._find_column(columns, ['kvantitet', 'antal'])
            grundtyp_col = self._find_column(columns, ['grundtyp'])  # New column for P51959
            
            # Parse using Excel outline levels
            parsed_items = []
            parent_stack = []  # Dynamic stack for parents at each level
            
            print(f"\nParsing {file_path.name}:")
            print(f"Total rows in DataFrame: {len(df)}")
            print(f"All columns in file: {columns}")
            print(f"Columns found: artikelnummer_col={artikelnummer_col}, benaming_col={benaming_col}, artikeltyp_col={artikeltyp_col}")
            
            # Parse directly from Excel to include ALL rows including operations
            print("\nParsing all Excel rows including operations:")
            
            # Get column indices from header
            header_row = 1
            col_indices = {}
            for col in range(1, ws.max_column + 1):
                header_val = ws.cell(row=header_row, column=col).value
                if header_val:
                    col_indices[header_val] = col
            
            artikel_col_idx = col_indices.get('Artikel/Operation', 3)
            benaming_col_idx = col_indices.get('Benämning', 4)
            artikeltyp_col_idx = col_indices.get('Artikeltyp/Operation', 1)
            grundtyp_col_idx = col_indices.get('Grundtyp', 2)
            kvantitet_col_idx = col_indices.get('Kvantitet', 5)
            
            # Parse each row from Excel - include ALL rows with data
            print(f"\nScanning all rows for data...")
            for excel_row in range(2, ws.max_row + 1):  # Skip header
                artikel_cell = ws.cell(row=excel_row, column=artikel_col_idx)
                
                # Check if row is hidden and log it
                row_dim = ws.row_dimensions.get(excel_row)
                is_hidden = row_dim.hidden if row_dim else False
                
                if artikel_cell.value is not None:
                    artikel = str(artikel_cell.value)
                    print(f"Processing row {excel_row}: {artikel} (hidden: {is_hidden})")
                    
                    # Get outline level directly from Excel grouping
                    row_dim = ws.row_dimensions.get(excel_row)
                    outline_level = row_dim.outline_level if row_dim else 0
                    
                    # Get values directly from Excel cells
                    benaming = ws.cell(row=excel_row, column=benaming_col_idx).value
                    artikeltyp = ws.cell(row=excel_row, column=artikeltyp_col_idx).value
                    grundtyp = ws.cell(row=excel_row, column=grundtyp_col_idx).value
                    kvantitet = ws.cell(row=excel_row, column=kvantitet_col_idx).value
                    
                    # Convert kvantitet to float if needed
                    if kvantitet is not None:
                        try:
                            kvantitet = float(kvantitet)
                        except:
                            kvantitet = None
                    
                    # Determine parent based on Excel outline level
                    parent_artikel = None
                    if outline_level > 0:
                        # Ensure parent_stack has enough elements
                        while len(parent_stack) <= (outline_level - 1):
                            parent_stack.append(None)
                        if outline_level - 1 < len(parent_stack):
                            parent_artikel = parent_stack[outline_level - 1]
                    
                    # Create item using Excel's grouping structure
                    item = TraceabilityItem(
                        artikelnummer=artikel,
                        artikelbenaming=str(benaming) if benaming else None,
                        artikeltyp=str(artikeltyp) if artikeltyp else None,
                        kvantitet=kvantitet,
                        grundtyp=str(grundtyp) if grundtyp and str(grundtyp).lower() not in ['nan', 'none', ''] else None,
                        parent_artikel=parent_artikel,
                        level=outline_level,
                        source_file=str(file_path),
                        source_type='nivålista'
                    )
                    
                    # Update parent stack for this level
                    while len(parent_stack) <= outline_level:
                        parent_stack.append(None)
                    parent_stack[outline_level] = artikel
                    for i in range(outline_level + 1, len(parent_stack)):
                        parent_stack[i] = None
                    
                    parsed_items.append(item)
                    items.append(item)
                    
                    
            
            print(f"\nParsed {len(parsed_items)} items from Excel (DataFrame had {len(df)} rows)")
            print(f"Total Excel data rows found: {data_rows}")
            level_counts = {}
            artikeltyp_counts = {}
            for item in parsed_items:
                level_counts[item.level] = level_counts.get(item.level, 0) + 1
                if item.artikeltyp:
                    artikeltyp_counts[item.artikeltyp] = artikeltyp_counts.get(item.artikeltyp, 0) + 1
            print(f"Level distribution: {level_counts}")
            print(f"Artikeltyp distribution: {artikeltyp_counts}")
            
            # Add all items to database with hierarchy in correct order
            for item in parsed_items:
                self.database.add_item_with_hierarchy(item)
                    
        except Exception as e:
            import traceback
            print(f"Error parsing {file_path}: {e}")
            print(f"Error type: {type(e).__name__}")
            print("Traceback:")
            traceback.print_exc()
        
        return items
    
    def _analyze_bom_structure(self, artikel: str, existing_items: List) -> tuple:
        """Analyze BOM structure to determine level and parent"""
        
        # Check if this is a sub-component (contains '/')
        if '/' in artikel:
            base_artikel = artikel.split('/')[0]
            # Find the base article in existing items
            for item in reversed(existing_items):  # Search from most recent
                if item.artikelnummer == base_artikel:
                    return item.level + 1, base_artikel
            # If base not found, assume it's level 1
            return 1, None
            
        # Check if this is a material/standard component (10-, 20-, 30-, 40-)
        if artikel.startswith(('10-', '20-', '30-', '40-')):
            # These are level 2 materials - find the most recent main assembly (level 0)
            # They belong to the assembly, not to individual KA components
            for item in reversed(existing_items):
                if item.level == 0:  # Main assembly
                    return 2, item.artikelnummer
            # If no assembly found, attach to most recent non-material item
            for item in reversed(existing_items):
                if not item.artikelnummer.startswith(('10-', '20-', '30-', '40-')):
                    return item.level + 1, item.artikelnummer
            return 2, None
            
        # Check if this is a purchased component (KA prefix)
        if artikel.startswith('KA'):
            # Find the most recent main assembly (level 0) - KA items belong to assemblies, not sub-parts
            for item in reversed(existing_items):
                if item.level == 0:  # Main assembly
                    return 1, item.artikelnummer
            # If no assembly found, make it level 1
            return 1, None
            
        # This is likely a main assembly
        return 0, None
    
    def _determine_hierarchy_level(self, artikel: str, parent_stack: List[str]) -> int:
        """Determine hierarchy level based on article naming pattern and BOM structure"""
        
        if '/' in artikel:
            # This is a sub-component (like 6192701676/2)
            base_artikel = artikel.split('/')[0]
            # Find the base article in parent stack and set level accordingly
            for i in range(len(parent_stack)):
                if parent_stack[i] == base_artikel or parent_stack[i].startswith(base_artikel):
                    return i + 1
            # If base article not found, it might be at top level with sub-components
            return 1
        elif artikel.startswith('KA'):
            # Purchased items (KA prefix) are usually components of the previous item
            return len(parent_stack)  # Same level as current context
        elif any(c in artikel for c in ['-', '_']) and not artikel.startswith(('10-', '20-', '30-', '40-')):
            # This looks like a main assembly (has dash but not material codes)
            return 0  # Top level
        else:
            # Material codes (10-, 20-, 30-, 40-) are purchased items, components of parent
            return len(parent_stack)  # Same level as current context
    
    def _update_parent_stack(self, artikel: str, level: int, parent_stack: List[str]):
        """Update the parent stack based on current article level"""
        # Only assemblies can be parents (not purchased items)
        is_assembly = (any(c in artikel for c in ['-', '_', '/']) and 
                      not artikel.startswith(('10-', '20-', '30-', '40-', 'KA')))
        
        if is_assembly:
            # Adjust stack size to current level
            while len(parent_stack) > level:
                parent_stack.pop()
            
            # Add current article as potential parent
            if level == len(parent_stack):
                parent_stack.append(artikel)
            elif level < len(parent_stack):
                parent_stack[level] = artikel
    
    def _parse_generic_file(self, file_path: Path) -> List[TraceabilityItem]:
        # Try to parse as a generic Excel file
        items = []
        try:
            df = pd.read_excel(file_path, engine='openpyxl' if file_path.suffix == '.xlsx' else 'xlrd')
            
            # Look for any columns that might contain relevant data
            columns = df.columns.tolist()
            artikelnummer_col = self._find_column(columns, ['artikelnummer', 'artikel', 'art.nr', 'article'])
            batch_col = self._find_column(columns, ['batchnummer', 'batch', 'serienummer', 'serial'])
            charge_col = self._find_column(columns, ['chargenummer', 'charge', 'lot'])
            
            if artikelnummer_col:
                for _, row in df.iterrows():
                    if pd.notna(row[artikelnummer_col]):
                        item = TraceabilityItem(
                            artikelnummer=str(row[artikelnummer_col]),
                            batchnummer=str(row[batch_col]) if batch_col and pd.notna(row[batch_col]) else None,
                            chargenummer=str(row[charge_col]) if charge_col and pd.notna(row[charge_col]) else None,
                            source_file=str(file_path),
                            source_type='generic'
                        )
                        items.append(item)
                        self.database.add_item(item)
        except Exception as e:
            print(f"Error parsing {file_path}: {e}")
        
        return items
    
    def _find_column(self, columns: List[str], search_terms: List[str]) -> Optional[str]:
        # Try exact matches first
        for term in search_terms:
            for col in columns:
                if term.lower() == str(col).lower():
                    return col
        
        # Then try partial matches
        for term in search_terms:
            for col in columns:
                col_lower = str(col).lower()
                if term.lower() in col_lower:
                    return col
        return None
    
    def get_database(self) -> TraceabilityDatabase:
        return self.database